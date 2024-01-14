from datetime import datetime
import openpyxl
from openpyxl.styles import Border, Side, PatternFill
import os.path

path = os.path.normpath("C:/Users/rshon/Documents/Etherbridge/Masterclass Trading Strategy/Masterclass portfolio.xlsx")
book = openpyxl.load_workbook(path.strip())

def values(Eth_Btc, Btc, Eth):
    Conseravtive_sheet = book["Conservative Trend"]

    # Clear all rows
    Conseravtive_sheet.delete_rows(1, Conseravtive_sheet.max_row)

    # Clear all columns
    Conseravtive_sheet.delete_cols(1, Conseravtive_sheet.max_column)

    Conseravtive_sheet["A1"] = "Ratio of ETH to BTC in conservative"
    Conseravtive_sheet["A2"] = "This is the probability of which major (Eth or BTC) is going to outperform the other."
    Conseravtive_sheet["A3"] = "Trend analysis performed on data series ETHBTC (I prefer Binance)"

    Conseravtive_sheet["B5"] = "ETHBTC Value"
    Conseravtive_sheet["C5"] = "ETHBTC Trend"
    Conseravtive_sheet["E5"] = "BTC Value"
    Conseravtive_sheet["F5"] = "BTC Trend"
    Conseravtive_sheet["H5"] = "ETH Value"
    Conseravtive_sheet["I5"] = "ETH Trend"

    Conseravtive_sheet["K5"] = "This means you'll want to hold:"
    Conseravtive_sheet["K6"] = "BTC"
    Conseravtive_sheet["K7"] = "ETH"

    count = 0
    ETH_BTC_sum = 0
    for list in Eth_Btc:
        Conseravtive_sheet["A"+ str(6+count)] = list[0]
        Conseravtive_sheet["B" + str(6 + count)] = list[1]
        Conseravtive_sheet["C" + str(6 + count)] = list[2]
        ETH_BTC_sum += list[2]
        count += 1

    ETH_BTC_count = count

    count = 0
    BTC_sum = 0
    for list in Btc:
        Conseravtive_sheet["E" + str(6 + count)] = list[1]
        cur = -1 if list[2] == 0 else 1
        Conseravtive_sheet["F" + str(6 + count)] = cur
        BTC_sum += cur
        count += 1

    BTC_TPI = float(BTC_sum)/count
    Conseravtive_sheet["F" + str(6 + count)] = BTC_TPI

    borders("F" + str(6 + count))
    highlight("F" + str(6 + count))

    count = 0
    ETH_sum = 0
    for list in Eth:
        Conseravtive_sheet["H" + str(6 + count)] = list[1]
        cur = -1 if list[2] == 0 else 1
        Conseravtive_sheet["I" + str(6 + count)] = cur
        ETH_sum += cur
        count += 1

    ETH_TPI = float(ETH_sum) / count
    Conseravtive_sheet["I" + str(6 + count)] = ETH_TPI

    borders("I" + str(6 + count))
    highlight("I" + str(6 + count))

    Conseravtive_sheet["A" + str(6 + count + 1)] = "BTC vs ETH TPI"
    Conseravtive_sheet["A" + str(6 + count + 2)] = "Seasonal Effect"
    Conseravtive_sheet["A" + str(6 + count + 3)] = "Relative Supply"

    # Get the current date
    current_date = datetime.now()
    # Check if it's between January 1st and June 1st
    seasonal_effect = 1 if (current_date.month >= 1 and current_date.month <= 6) else 0
    ETH_BTC_sum += seasonal_effect
    ETH_BTC_count += 1

    if ETH_TPI > BTC_TPI:
        BTC_vs_ETH_TPI = 1
        ETH_BTC_sum += 1
        ETH_BTC_count += 1
    elif ETH_TPI == BTC_TPI:
        BTC_vs_ETH_TPI = "null"
    else:
        BTC_vs_ETH_TPI = 0
        ETH_BTC_count += 1

    # Relative supply
    ETH_BTC_sum += 1
    ETH_BTC_count += 1

    Conseravtive_sheet["C" + str(6 + count + 1)] = BTC_vs_ETH_TPI
    Conseravtive_sheet["C" + str(6 + count + 2)] = seasonal_effect
    Conseravtive_sheet["C" + str(6 + count + 3)] = 1
    Conseravtive_sheet["C" + str(6 + count + 4)] = float(ETH_BTC_sum)/ETH_BTC_count

    borders("C" + str(6 + count + 4))
    highlight("C" + str(6 + count + 4))

    Conseravtive_sheet["L6"] = 1 - (float(ETH_BTC_sum)/ETH_BTC_count)
    Conseravtive_sheet["L7"] = float(ETH_BTC_sum)/ETH_BTC_count

    borders("L6")
    highlight("L6")
    borders("L7")
    highlight("L7")

    Conseravtive_sheet["K9"] = "BTC TPI"
    Conseravtive_sheet["K10"] = "ETH TPI"
    Conseravtive_sheet["L9"] = BTC_TPI
    Conseravtive_sheet["L10"] = ETH_TPI

    borders("L9")
    highlight("L9")
    borders("L10")
    highlight("L10")

    book.save("Masterclass portfolio.xlsx")

def borders(cell):
    Conseravtive_sheet = book["Conservative Trend"]
    cell = Conseravtive_sheet[cell]

    # Create a border style
    border_style = Border(left=Side(border_style='thin', color='000000'),
                          right=Side(border_style='thin', color='000000'),
                          top=Side(border_style='thin', color='000000'),
                          bottom=Side(border_style='thin', color='000000'))

    # Apply the border style to the cell
    cell.border = border_style

def highlight(cell):
    Conseravtive_sheet = book["Conservative Trend"]
    cell = Conseravtive_sheet[cell]

    # Create a fill style (e.g., yellow fill)
    fill_style = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Apply the fill style to the cell
    cell.fill = fill_style